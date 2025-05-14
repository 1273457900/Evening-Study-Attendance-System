import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from models import User, AttendanceRecord, db, ClassroomAdmin, LeaveRecord, AbsenceRecord
from werkzeug.security import generate_password_hash
from flask import current_app
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage, ImageDraw, ImageFont
import textwrap

def generate_template():
    """生成Excel模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生名单模板"
    
    # 设置标题行
    headers = ["班级", "学号", "姓名", "教室位置"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=1, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    
    # 添加一些示例数据
    example_data = [
        ["物联网2321", "20230101", "张三",  "G503"],
        ["物联网2321", "20230102", "李四",  "G503"],
        ["物联网2321", "20230103", "王五",  "G503"]
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # 添加说明
    ws.cell(row= 3, column=6, value="说明：")
    ws.cell(row= 3, column=6).font = Font(bold=True)
    ws.cell(row= 4, column=6, value="1. 班级: 请填写学生所在班级")
    ws.cell(row= 5, column=6, value="2. 学号: 请填写学生学号，将作为登录用户名")
    ws.cell(row= 6, column=6, value="3. 姓名: 请填写学生姓名")

    ws.cell(row= 8, column=6, value="5. 教室位置: 学生所在教室编号（如G503），用于关联到教室管理员")
    
    return wb

def import_students_from_excel(file_path):
    """从Excel文件导入学生名单"""
    try:
        df = pd.read_excel(file_path, usecols="A:E")
        
        # 验证数据格式
        required_columns = ["班级", "学号", "姓名"]
        for col in required_columns:
            if col not in df.columns:
                return False, f"Excel文件格式错误，缺少'{col}'列"
        
        # 开始导入学生数据
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        for _, row in df.iterrows():
            class_name = str(row["班级"])
            student_id = str(row["学号"])
            name = str(row["姓名"])
            
            # 处理学号 - 如果是浮点数格式（如20230101.0），转换为整数
            try:
                if '.' in student_id:  # 检查是否是浮点数格式
                    student_id = str(int(float(student_id)))
                else:
                    student_id = str(int(student_id))  # 确保是整数格式
            except ValueError:
                skipped_count += 1
                continue  # 如果学号无法转换为数字，跳过该学生
            
            # 检查是否已存在该学生
            existing_user = User.query.filter_by(username=student_id).first()
            
            if existing_user:
                # 更新已存在的学生信息
                existing_user.name = name
                existing_user.class_name = class_name
                if "教室位置" in row and not pd.isna(row["教室位置"]):
                    existing_user.classroom_location = str(row["教室位置"])
                updated_count += 1
            else:
                # 创建新学生 - 不设置密码
                new_user = User(
                    username=student_id,
                    name=name,
                    class_name=class_name,
                    role="student"
                )
                if "教室位置" in row and not pd.isna(row["教室位置"]):
                    new_user.classroom_location = str(row["教室位置"])
                db.session.add(new_user)
                imported_count += 1
        
        db.session.commit()
        return True, f"成功导入{imported_count}名新学生，更新{updated_count}名已有学生信息，跳过{skipped_count}名格式错误的学生"
    
    except Exception as e:
        db.session.rollback()
        return False, f"导入失败: {str(e)}"

def export_attendance_records(start_date, end_date, class_name=None, export_photos=True):
    """导出考勤记录到Excel，并根据选项导出相关照片"""
    try:
        # 查询考勤记录
        query = db.session.query(
            User.username, 
            db.func.coalesce(AttendanceRecord.student_name, User.name).label('name'), 
            db.func.coalesce(AttendanceRecord.class_name, User.class_name).label('class_name'),
            AttendanceRecord.date, 
            AttendanceRecord.sign_in_time,
            AttendanceRecord.leave_time,
            AttendanceRecord.return_time,
            AttendanceRecord.id.label('attendance_id'),
            User.id.label('student_id'),
            AttendanceRecord.sign_in_photo,
            AttendanceRecord.leave_photo,
            AttendanceRecord.return_photo
        ).join(
            AttendanceRecord, User.id == AttendanceRecord.student_id
        ).filter(
            AttendanceRecord.date.between(start_date, end_date),
            User.role == "student"
        )
        
        # 如果指定了班级，添加筛选条件
        if class_name:
            query = query.filter(User.class_name == class_name)
            
        # 按日期和班级排序
        records = query.order_by(AttendanceRecord.date, User.class_name, User.username).all()
        
        if not records:
            return None
        
        # 生成导出基础文件名（用于Excel和照片文件夹）
        if class_name:
            base_filename = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}-{class_name}-考勤记录"
        else:
            base_filename = f"{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}-全部班级-考勤记录"
        
        # 导出路径
        export_base_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'exports')
        export_excel_path = os.path.join(export_base_dir, f"{base_filename}.xlsx")
        export_photos_dir = os.path.join(export_base_dir, f"{base_filename}_照片")
        
        # 确保导出目录存在
        os.makedirs(export_base_dir, exist_ok=True)
        
        # 如果需要导出照片，创建照片目录
        if export_photos:
            os.makedirs(export_photos_dir, exist_ok=True)
            
        # 创建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "考勤记录"
        
        # 设置表头 - 根据是否导出照片决定是否包含照片列
        base_headers = ["日期", "班级", "学号", "姓名", "签到时间", "暂离时间", "返回时间", "今日暂离次数", "今日暂离分钟数"]
        headers = base_headers.copy()
        
        # 添加照片列标题
        if export_photos:
            headers.append("签到照片")
            headers.append("暂离/返回照片")
            
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 调整列宽
        for col in range(1, len(base_headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 如果导出照片，设置照片列的宽度和行高
        if export_photos:
            # 照片列宽度设置宽一些
            ws.column_dimensions[openpyxl.utils.get_column_letter(len(base_headers) + 1)].width = 20
            ws.column_dimensions[openpyxl.utils.get_column_letter(len(base_headers) + 2)].width = 25
            
        # 写入数据并导出照片
        row = 2
        for record in records:
            student_folder_name = ""
            photo_paths = []
            
            # 如果需要导出照片，处理照片缓存
            if export_photos:
                # 为每个学生创建照片文件夹
                date_str = record.date.strftime('%Y%m%d')
                student_folder_name = f"{date_str}_{record.class_name}_{record.username}_{record.name}"
                student_photo_dir = os.path.join(export_photos_dir, student_folder_name)
                os.makedirs(student_photo_dir, exist_ok=True)
                
                # 设置行高（照片行需要更高一些）
                ws.row_dimensions[row].height = 100
            
            # 查询该学生的所有暂离记录
            leave_records = LeaveRecord.query.filter(
                LeaveRecord.student_id == record.student_id,
                LeaveRecord.attendance_id == record.attendance_id
            ).order_by(LeaveRecord.leave_time).all()
            
            # 计算暂离次数和分钟数
            leave_count = len(leave_records)
            total_minutes = 0
            
            # 处理暂离记录
            leave_return_photos = []
            for i, leave in enumerate(leave_records, 1):
                # 计算暂离时长
                if leave.leave_time and leave.return_time:
                    duration = (leave.return_time - leave.leave_time).total_seconds() / 60
                    total_minutes += duration
                
                # 如果需要导出照片，收集暂离和返回照片
                if export_photos:
                    # 处理暂离照片
                    if leave.leave_photo:
                        src_path = os.path.join(current_app.config['UPLOAD_FOLDER'], leave.leave_photo)
                        if os.path.exists(src_path):
                            # 保存照片文件到文件夹
                            dest_filename = f"暂离{i}_{leave.leave_time.strftime('%H%M%S')}.png" if leave.leave_time else f"暂离{i}.png"
                            dest_path = os.path.join(student_photo_dir, dest_filename)
                            try:
                                import shutil
                                shutil.copy2(src_path, dest_path)
                            except Exception as e:
                                current_app.logger.error(f"复制暂离照片失败: {str(e)}")
                            
                            # 添加到要在Excel中显示的照片集合
                            leave_return_photos.append({
                                'src_path': src_path,
                                'type': 'leave',
                                'index': i,
                                'time': leave.leave_time
                            })
                    
                    # 处理返回照片
                    if leave.return_photo:
                        src_path = os.path.join(current_app.config['UPLOAD_FOLDER'], leave.return_photo)
                        if os.path.exists(src_path):
                            # 保存照片文件到文件夹
                            dest_filename = f"返回{i}_{leave.return_time.strftime('%H%M%S')}.png" if leave.return_time else f"返回{i}.png"
                            dest_path = os.path.join(student_photo_dir, dest_filename)
                            try:
                                import shutil
                                shutil.copy2(src_path, dest_path)
                            except Exception as e:
                                current_app.logger.error(f"复制返回照片失败: {str(e)}")
                            
                            # 添加到要在Excel中显示的照片集合
                            leave_return_photos.append({
                                'src_path': src_path,
                                'type': 'return',
                                'index': i,
                                'time': leave.return_time
                            })
            
            # 填充数据
            ws.cell(row=row, column=1).value = record.date.strftime('%Y-%m-%d')
            ws.cell(row=row, column=2).value = record.class_name
            ws.cell(row=row, column=3).value = record.username
            ws.cell(row=row, column=4).value = record.name
            ws.cell(row=row, column=5).value = record.sign_in_time.strftime('%H:%M:%S') if record.sign_in_time else "未签到"
            ws.cell(row=row, column=6).value = record.leave_time.strftime('%H:%M:%S') if record.leave_time else "-"
            ws.cell(row=row, column=7).value = record.return_time.strftime('%H:%M:%S') if record.return_time else "-"
            ws.cell(row=row, column=8).value = leave_count
            ws.cell(row=row, column=9).value = round(total_minutes)
            
            # 如果导出照片，插入照片到Excel单元格
            if export_photos:
                # 处理签到照片
                if record.sign_in_photo:
                    sign_in_path = os.path.join(current_app.config['UPLOAD_FOLDER'], record.sign_in_photo)
                    if os.path.exists(sign_in_path):
                        # 保存照片文件到文件夹
                        dest_filename = f"签到_{record.sign_in_time.strftime('%H%M%S')}.png" if record.sign_in_time else "签到.png"
                        dest_path = os.path.join(student_photo_dir, dest_filename)
                        try:
                            import shutil
                            shutil.copy2(sign_in_path, dest_path)
                        except Exception as e:
                            current_app.logger.error(f"复制签到照片失败: {str(e)}")
                        
                        try:
                            # 使用PIL调整图片大小（照片在拍摄时已添加水印）
                            img = PILImage.open(sign_in_path)
                            
                            # 调整大小用于嵌入Excel，使用高质量调整
                            width_height = max(150, int(min(img.width, img.height) / 4))  # 更大的尺寸
                            img_resized = img.resize((width_height, width_height), PILImage.LANCZOS)
                            
                            # 保存调整后的图片到临时文件，使用高质量设置
                            temp_path = os.path.join(student_photo_dir, "temp_sign_in.png")
                            img_resized.save(temp_path, quality=100, optimize=True)
                            
                            # 将图片插入到Excel
                            img = Image(temp_path)
                            cell_address = f"{openpyxl.utils.get_column_letter(len(base_headers) + 1)}{row}"
                            ws.add_image(img, cell_address)
                        except Exception as e:
                            current_app.logger.error(f"插入签到照片失败: {str(e)}")
                            ws.cell(row=row, column=len(base_headers) + 1).value = "照片处理失败"
                
                # 处理暂离/返回照片组合
                if leave_return_photos:
                    try:
                        # 创建组合图片（最多显示3个暂离记录的照片）
                        max_display = min(3, len(leave_return_photos))
                        displayed_photos = leave_return_photos[:max_display]
                        
                        # 计算组合图片大小
                        photo_size = 150  # 单个照片尺寸
                        combined_width = photo_size * max_display
                        combined_height = photo_size
                        
                        # 创建空白画布
                        combined_img = PILImage.new('RGB', (combined_width, combined_height), color=(255, 255, 255))
                        
                        # 添加每张照片到画布
                        for idx, photo_info in enumerate(displayed_photos):
                            try:
                                # 照片在拍摄时已添加水印，只需调整大小
                                img = PILImage.open(photo_info['src_path'])
                                img_resized = img.resize((photo_size, photo_size), PILImage.LANCZOS)
                                combined_img.paste(img_resized, (idx * photo_size, 0))
                                
                                # 添加标签（暂离/返回）
                                draw = ImageDraw.Draw(combined_img)
                                label = f"{photo_info['type']}_{photo_info['index']}"
                                # 尝试获取默认字体，如果失败则忽略文字绘制
                                try:
                                    # 在Windows上可能需要系统字体路径
                                    font_path = "static/fonts/SourceHanSansSC-Regular.otf"
                                    font = ImageFont.truetype(font_path, 14)
                                    draw.text((idx * photo_size + 5, 5), label, fill="white", font=font)
                                except:
                                    pass  # 字体加载失败时忽略文字绘制
                            except Exception as e:
                                current_app.logger.error(f"处理单个照片失败: {str(e)}")
                                continue
                        
                        # 保存组合图片
                        temp_path = os.path.join(student_photo_dir, "temp_combined.png")
                        combined_img.save(temp_path, quality=95, optimize=True)
                        
                        # 将组合图片插入到Excel
                        img = Image(temp_path)
                        cell_address = f"{openpyxl.utils.get_column_letter(len(base_headers) + 2)}{row}"
                        ws.add_image(img, cell_address)
                        
                        # 如果有更多照片未显示，在单元格中添加说明
                        if len(leave_return_photos) > max_display:
                            cell = ws.cell(row=row, column=len(base_headers) + 2)
                            cell.comment = openpyxl.comments.Comment(
                                f"还有{len(leave_return_photos) - max_display}张照片未显示，请查看照片文件夹", "系统"
                            )
                    except Exception as e:
                        current_app.logger.error(f"创建暂离/返回照片组合失败: {str(e)}")
                        ws.cell(row=row, column=len(base_headers) + 2).value = "照片处理失败"
                        
            row += 1
        
        # 保存Excel文件
        wb.save(export_excel_path)
        
        # 返回导出文件夹名称（包含Excel和照片文件夹的共同父文件夹）
        return f"{base_filename}.xlsx"
    except Exception as e:
        current_app.logger.error(f"导出考勤记录失败: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        return None

def generate_admin_template():
    """生成教室管理员账号导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教室管理员账号"
    
    # 设置表头
    headers = ['教室名称', '登录账号', '初始密码', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 设置示例数据
    example_data = [
        ['G503', 'g503', '123456', '示例数据，导入时请删除'],
        ['G504', 'g504', '123456', '密码不填则默认为123456'],
        ['G505', 'g505_admin', '123456', '账号可自定义，不必与教室名称完全一致'],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
    
    # 添加说明
    notes_row = len(example_data) + 3
    ws.cell(row=notes_row, column=1, value="说明：").font = Font(bold=True)
    ws.cell(row=notes_row+1, column=1, value="1. 教室名称: 请填写教室的标准名称，如G503")
    ws.cell(row=notes_row+2, column=1, value="2. 登录账号: 可以是教室名称的小写或其他自定义账号")
    ws.cell(row=notes_row+3, column=1, value="3. 初始密码: 不填则默认为123456")
    ws.cell(row=notes_row+4, column=1, value="4. 学生将根据'教室位置'字段关联到对应教室管理员")
    ws.cell(row=notes_row+5, column=1, value="5. 请确保教室名称与学生导入模板中的'教室位置'字段匹配")
    
    # 调整列宽
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20
    
    return wb

def import_admin_accounts(file_path):
    """从Excel文件导入教室管理员账号"""
    try:
        df = pd.read_excel(file_path)
        required_columns = ['教室名称', '登录账号']
        
        # 检查必要列是否存在
        if not all(col in df.columns for col in required_columns):
            return False, "Excel文件格式错误，请使用正确的模板"
        
        success_count = 0
        error_messages = []
        
        for index, row in df.iterrows():
            try:
                # 跳过空行
                if pd.isna(row['教室名称']) or pd.isna(row['登录账号']):
                    continue
                    
                classroom_name = str(row['教室名称']).strip().upper()
                username = str(row['登录账号']).strip()
                password = str(row['初始密码']).strip() if not pd.isna(row['初始密码']) else '123456'
                
                # 检查账号是否已存在
                existing_admin = ClassroomAdmin.query.filter_by(username=username).first()
                if existing_admin:
                    error_messages.append(f"账号 {username} 已存在")
                    continue
                
                # 创建新账号
                admin = ClassroomAdmin(
                    username=username,
                    classroom_name=classroom_name,
                )
                admin.set_password(password)  # 此方法会同时设置密码哈希和明文
                db.session.add(admin)
                db.session.commit()  # 先提交以获取admin.id
                
                # 关联已有学生
                students = User.query.filter_by(classroom_location=classroom_name).all()
                for student in students:
                    student.classroom_id = admin.id
                
                success_count += 1
                
            except Exception as e:
                error_messages.append(f"第 {index+2} 行导入失败：{str(e)}")
        
        db.session.commit()
        
        # 生成结果消息
        message = f"成功导入 {success_count} 个账号。"
        if error_messages:
            message += f"\n导入错误：\n" + "\n".join(error_messages)
        
        return True, message
        
    except Exception as e:
        return False, f"导入失败：{str(e)}"

def export_admin_accounts():
    """导出教室管理员账号列表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教室管理员账号"
    
    # 设置表头
    headers = ['教室名称', '登录账号', '密码', '状态', '创建时间', '最后登录时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 获取所有账号
    admins = ClassroomAdmin.query.order_by(ClassroomAdmin.classroom_name).all()
    
    # 写入数据
    for row_idx, admin in enumerate(admins, 2):
        ws.cell(row=row_idx, column=1, value=admin.classroom_name)
        ws.cell(row=row_idx, column=2, value=admin.username)
        ws.cell(row=row_idx, column=3, value=admin.password_text or '******')
        ws.cell(row=row_idx, column=4, value='启用' if admin.is_active else '禁用')
        ws.cell(row=row_idx, column=5, value=admin.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=6, value=admin.last_login.strftime('%Y-%m-%d %H:%M:%S') if admin.last_login else '未登录')
    
    # 调整列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20
    
    return wb

def add_timestamp_watermark(img, timestamp_text, position='center', no_background=False):
    """给图片添加固定大小、明显清晰的水印（完全居中显示）"""
    img_with_watermark = img.copy()
    draw = ImageDraw.Draw(img_with_watermark)
    
    # 动态调整字体大小（根据图片高度）
    font_size = max(24, min(48, int(img.height / 20)))  # 在24-48px之间，按比例调整
    
    try:
        font = ImageFont.truetype("static/fonts/SourceHanSansSC-Regular.otf", font_size)
    except:
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()
    
    # 计算多行文本高度和每行宽度
    lines = timestamp_text.split('\n')
    line_height = font_size + 5
    total_text_height = len(lines) * line_height
    line_widths = [font.getlength(line) for line in lines]
    max_text_width = max(line_widths) if line_widths else 0
    
    # 计算居中位置（确保不会超出边界）
    margin = 30
    if position == 'center':
        x = (img.width - max_text_width) // 2
        y = (img.height - total_text_height) // 2
    elif position == 'bottom':
        x = (img.width - max_text_width) // 2
        y = img.height - total_text_height - margin
    else:  # 默认居中
        x = (img.width - max_text_width) // 2
        y = (img.height - total_text_height) // 2
    
    # 添加半透明背景（确保能容纳多行文本）
    if not no_background:
        bg_margin = 10
        draw.rectangle(
            [x - bg_margin, y - bg_margin, 
             x + max_text_width + bg_margin, y + total_text_height + bg_margin],
            fill=(0, 0, 0, 180)
        )
    
    # 绘制多行文本（每行单独计算水平居中）
    for i, line in enumerate(lines):
        line_x = (img.width - line_widths[i]) // 2  # 每行单独计算水平居中
        line_y = y + i * line_height
        
        # 红色描边
        for offset in range(-2, 3):
            draw.text((line_x + offset, line_y), line, fill="red", font=font)
            draw.text((line_x, line_y + offset), line, fill="red", font=font)
        
        # 白色主文字
        draw.text((line_x, line_y), line, fill="white", font=font)
    
    return img_with_watermark

def delete_student(student_id):
    """仅删除学生记录（保留所有关联记录）"""
    try:
        student = User.query.get(student_id)
        if not student:
            return False, "学生不存在"
        
        # 1. 不删除任何关联记录
        # AttendanceRecord.query.filter_by(student_id=student_id).delete()
        # LeaveRecord.query.filter_by(student_id=student_id).delete()
        # AbsenceRecord.query.filter_by(student_id=student_id).delete()
        
        # 2. 仅删除学生本身
        db.session.delete(student)
        db.session.commit()
        
        return True, "学生删除成功（所有关联记录已保留）"
    except Exception as e:
        db.session.rollback()
        return False, f"删除失败：{str(e)}" 