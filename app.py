from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, send_from_directory
import os
import socket
import logging
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from models import db, User, AttendanceRecord, ClassroomAdmin, LeaveRecord, AbsenceRecord
from forms import LoginForm, ImportStudentsForm, ExportAttendanceForm
from utils import generate_template, import_students_from_excel, export_attendance_records, generate_admin_template, import_admin_accounts, export_admin_accounts, add_timestamp_watermark
from generate_cert import generate_ssl_cert
from functools import wraps

# 配置日志
logging.basicConfig(level=logging.DEBUG, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__, 
           static_url_path='/static',  # 确保这个与模板中的引用一致
           static_folder='static')  # 确保这是相对于app.py的正确路径
app.config['SECRET_KEY'] = 'your-secret-key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///attendance.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB最大上传限制

# 确保上传目录存在
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# 初始化数据库
db.init_app(app)

# 初始化登录管理器
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '请先登录再访问此页面'

@login_manager.user_loader
def load_user(user_id):
    # 尝试查找用户
    try:
        # 先检查是否是普通用户ID
        user_type, user_id = user_id.split('_')
        user_id = int(user_id)
        
        if user_type == 'user':
            return User.query.get(user_id)
        elif user_type == 'admin':
            return ClassroomAdmin.query.get(user_id)
    except ValueError:
        # 如果不是新格式，尝试兼容旧格式（纯数字ID）
        try:
            return User.query.get(int(user_id))
        except:
            logger.error(f"无法加载用户ID: {user_id}")
            return None
    except Exception as e:
        logger.error(f"加载用户时出错: {str(e)}")
        return None

# 添加上下文处理器，为所有模板提供now变量
@app.context_processor
def inject_now():
    return {'now': datetime.now()}

@app.context_processor
def inject_get_today_record():
    """提供一个函数用于获取学生当天的签到记录"""
    def get_today_record(student_id):
        today = datetime.now().date()
        attendance = AttendanceRecord.query.filter_by(
            student_id=student_id,
            date=today
        ).first()
        
        if attendance:
            # 查询最新的暂离记录
            latest_leave = LeaveRecord.query.filter(
                LeaveRecord.student_id == student_id,
                LeaveRecord.attendance_id == attendance.id
            ).order_by(LeaveRecord.leave_time.desc()).first()
            
            # 如果存在最新的暂离记录且未返回，则更新attendance的暂离状态
            if latest_leave and latest_leave.is_active:
                attendance.leave_time = latest_leave.leave_time
                attendance.return_time = None
            # 如果最新记录已返回，则更新返回时间
            elif latest_leave and latest_leave.return_time:
                attendance.leave_time = latest_leave.leave_time
                attendance.return_time = latest_leave.return_time
        
        return attendance
    return {'get_today_record': get_today_record}

@app.context_processor
def inject_get_today_absence():
    """提供一个函数用于获取学生当天的请假记录"""
    def get_today_absence(student_id):
        today = datetime.now().date()
        # 查询当天的请假记录
        absence = AbsenceRecord.query.filter_by(
            student_id=student_id,
            date=today
        ).first()
        
        return absence
    return {'get_today_absence': get_today_absence}

@app.context_processor
def inject_get_active_leave_record():
    """提供一个函数用于获取学生的活跃暂离记录"""
    def get_active_leave_record(student_id, attendance_id):
        # 查询未返回的暂离记录
        return LeaveRecord.query.filter(
            LeaveRecord.student_id == student_id,
            LeaveRecord.attendance_id == attendance_id,
            LeaveRecord.return_time.is_(None)
        ).order_by(LeaveRecord.leave_time.desc()).first()
    return {'get_active_leave_record': get_active_leave_record}

@app.context_processor
def inject_get_latest_leave_record():
    """提供一个函数用于获取学生的最新暂离记录(已返回的也包括)"""
    def get_latest_leave_record(student_id, attendance_id):
        # 查询最新的暂离记录（无论是否返回）
        return LeaveRecord.query.filter(
            LeaveRecord.student_id == student_id,
            LeaveRecord.attendance_id == attendance_id
        ).order_by(LeaveRecord.leave_time.desc()).first()
    return {'get_latest_leave_record': get_latest_leave_record}

@app.context_processor
def inject_get_leave_count_and_minutes():
    """提供函数用于计算学生的暂离次数和暂离分钟数"""
    def get_leave_count(student_id, attendance_id):
        """获取学生当天的暂离次数"""
        count = LeaveRecord.query.filter(
            LeaveRecord.student_id == student_id,
            LeaveRecord.attendance_id == attendance_id
        ).count()
        return count or 0
    
    def get_leave_minutes(student_id, attendance_id):
        """获取学生当天的暂离总分钟数"""
        leave_records = LeaveRecord.query.filter(
            LeaveRecord.student_id == student_id,
            LeaveRecord.attendance_id == attendance_id
        ).all()
        
        total_minutes = 0
        for record in leave_records:
            if record.leave_time:
                # 如果已返回，计算实际暂离时间
                if record.return_time:
                    duration = (record.return_time - record.leave_time).total_seconds() / 60
                # 如果未返回，计算到当前时间的暂离时间
                else:
                    duration = (datetime.now() - record.leave_time).total_seconds() / 60
                total_minutes += duration
        
        return round(total_minutes)
    
    return {
        'get_leave_count': get_leave_count,
        'get_leave_minutes': get_leave_minutes
    }

def get_local_ip():
    """获取本机局域网IP地址"""
    try:
        # 创建一个临时socket连接到外部，获取本机在局域网中的IP
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        # 不需要真正连接到这个地址，只是让系统选择网络接口
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        # 如果上面的方法失败，尝试获取所有可能的IP
        try:
            host_name = socket.gethostname()
            ip_list = socket.gethostbyname_ex(host_name)[2]
            # 过滤出非localhost的IP
            ip_list = [ip for ip in ip_list if not ip.startswith("127.")]
            if ip_list:
                return ip_list[0]
        except:
            pass
        return "127.0.0.1"  # 默认返回localhost

@app.route('/')
def index():
    if current_user.is_authenticated:
        logger.debug(f"已登录用户: {current_user}, 类型: {type(current_user)}")
        if isinstance(current_user, User):
            if current_user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        elif isinstance(current_user, ClassroomAdmin):
            return redirect(url_for('classroom_admin_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        logger.debug(f"尝试登录用户名: {username}")
        
        # 先尝试作为普通用户登录
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            flash('登录成功！', 'success')
            logger.debug(f"普通用户登录成功: {username}")
            if user.role == 'teacher':
                return redirect(url_for('teacher_dashboard'))
            else:
                return redirect(url_for('student_dashboard'))
        else:
            # 尝试作为教室管理员登录
            admin = ClassroomAdmin.query.filter(db.func.lower(ClassroomAdmin.username) == username.lower()).first()
            if admin and admin.check_password(password):
                if not admin.is_active:
                    flash('账号已禁用，请联系管理员', 'danger')
                    logger.warning(f"禁用账号尝试登录: {username}")
                    return render_template('login.html', form=form)
                
                # 记录最后登录时间
                admin.last_login = datetime.now()
                db.session.commit()
                
                # 登录管理员账号
                login_user(admin)
                flash('教室管理员登录成功！', 'success')
                logger.debug(f"教室管理员登录成功: {username}")
                return redirect(url_for('classroom_admin_dashboard'))
            else:
                flash('用户名或密码错误', 'danger')
                logger.warning(f"登录失败，用户名或密码错误: {username}")
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/student')
@login_required
def student_dashboard():
    # 检查是否是学生用户
    if not isinstance(current_user, User) or current_user.role != 'student':
        flash('无权访问学生页面', 'danger')
        logger.warning(f"非学生用户尝试访问学生页面: {current_user}")
        return redirect(url_for('index'))
    
    # 获取今天的考勤记录
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=current_user.id,
        date=today
    ).first()
    
    return render_template('student_dashboard.html', attendance=attendance)

@app.route('/teacher')
@login_required
def teacher_dashboard():
    # 检查是否是教师用户
    if not isinstance(current_user, User) or current_user.role != 'teacher':
        flash('无权访问教师页面', 'danger')
        logger.warning(f"非教师用户尝试访问教师页面: {current_user}")
        return redirect(url_for('index'))
    
    # 获取今天的考勤统计
    today = datetime.now().date()
    
    # 查询所有学生
    students = User.query.filter_by(role='student').all()
    
    # 查询今天已签到的学生
    signed_in = db.session.query(User).join(
        AttendanceRecord, User.id == AttendanceRecord.student_id
    ).filter(
        AttendanceRecord.date == today,
        AttendanceRecord.sign_in_time.isnot(None)
    ).count()
    
    # 计算签到率
    total_students = len(students)
    signed_in_rate = round(signed_in / total_students * 100, 2) if total_students > 0 else 0
    
    return render_template('teacher_dashboard.html', 
                         total_students=total_students,
                         signed_in=signed_in,
                         signed_in_rate=signed_in_rate)

@app.route('/camera')
@login_required
def camera():
    """相机拍照页面"""
    # 检查用户角色
    action = request.args.get('action', 'sign_in')
    student_id = request.args.get('student_id')
    
    if isinstance(current_user, ClassroomAdmin):
        # 教室管理员
        # 获取该教室的学生列表（通过教室位置与管理员用户名匹配）
        students = User.query.filter(
            User.role == 'student',
            db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
        ).all()
        
        # 如果有指定学生，则预先选择该学生
        selected_student = None
        if student_id:
            selected_student = User.query.filter(
                User.id == student_id, 
                User.role == 'student',
                db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
            ).first()
            if not selected_student:
                flash('指定的学生不存在或不属于您管理的教室', 'warning')
        
        return render_template('camera.html', 
                               action=action, 
                               admin_id=current_user.id,
                               students=students,
                               selected_student=selected_student)
    elif current_user.role == 'student':
        # 学生
        # 目前学生只能暂离和返回拍照
        if action not in ['leave', 'return']:
            flash('学生不能自己签到，请联系教室管理员', 'warning')
            return redirect(url_for('student_dashboard'))
        return render_template('camera.html', action=action)
    else:
        # 其他角色不允许使用拍照功能
        flash('您没有权限使用拍照功能', 'danger')
        return redirect(url_for('index'))

@app.route('/admin_photo', methods=['POST'])
@login_required
def admin_photo():
    """教室管理员拍照处理"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以拍照'})
    
    # 获取POST数据
    photo_data = request.form.get('photo')
    student_id = request.form.get('student_id')
    action = request.form.get('action', 'sign_in')
    watermark_position = request.form.get('watermark_position', 'bottom')
    
    if not photo_data:
        return jsonify({'success': False, 'message': '没有收到照片数据'})
        
    if not student_id:
        return jsonify({'success': False, 'message': '没有指定学生'})
    
    # 查询该学生是否属于该教室管理员（通过教室位置与管理员用户名匹配）
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 保存照片
    photo_path = save_photo(photo_data, f"{action}_{student_id}")
    
    # 记录签到/暂离/返回信息
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance:
        # 如果没有签到记录，先创建一个
        attendance = AttendanceRecord(
            student_id=student_id,
            student_name=student.name,
            class_name=student.class_name,
            date=today
        )
        db.session.add(attendance)
    
    # 根据action类型更新不同的记录
    if action == 'sign_in':
        attendance.sign_in_time = datetime.now()
        attendance.sign_in_photo = photo_path
        message = '签到成功'
    elif action == 'leave':
        attendance.leave_time = datetime.now()
        attendance.leave_photo = photo_path
        message = '暂离已记录'
    elif action == 'return':
        attendance.return_time = datetime.now()
        attendance.return_photo = photo_path
        message = '返回已记录'
    else:
        return jsonify({'success': False, 'message': '未知操作类型'})
    
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': message,
        'student_name': student.name,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/leave', methods=['POST'])
@login_required
def leave():
    """学生暂离处理"""
    if current_user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以登记暂离'})
    
    # 获取POST数据
    photo_data = request.form.get('photo')
    watermark_position = request.form.get('watermark_position', 'bottom')
    
    if not photo_data:
        return jsonify({'success': False, 'message': '没有收到照片数据'})
    
    # 保存照片
    photo_path = save_photo(photo_data, f"leave_{current_user.id}")
    
    # 查找今天的考勤记录
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=current_user.id,
        date=today
    ).first()
    
    if not attendance:
        # 如果没有签到记录，先创建一个
        attendance = AttendanceRecord(
            student_id=current_user.id,
            student_name=current_user.name,
            class_name=current_user.class_name,
            date=today
        )
        db.session.add(attendance)
        db.session.commit()  # 提交以获取ID
    
    # 检查是否有未返回的暂离记录
    active_leave = LeaveRecord.query.filter(
        LeaveRecord.student_id == current_user.id,
        LeaveRecord.attendance_id == attendance.id,
        LeaveRecord.return_time.is_(None)
    ).first()
    
    if active_leave:
        return jsonify({'success': False, 'message': '您有未完成的暂离记录，请先登记返回'})
    
    # 创建新的暂离记录
    leave_record = LeaveRecord(
        student_id=current_user.id,
        student_name=current_user.name,
        class_name=current_user.class_name,
        attendance_id=attendance.id,
        leave_time=datetime.now(),
        leave_photo=photo_path
    )
    
    db.session.add(leave_record)
    
    # 保持向后兼容性
    attendance.leave_time = datetime.now()
    attendance.leave_photo = photo_path
    
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': '暂离已记录',
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/return_signin', methods=['POST'])
@login_required
def return_signin():
    """学生返回签到处理"""
    if current_user.role != 'student':
        return jsonify({'success': False, 'message': '只有学生可以登记返回'})
    
    # 获取POST数据
    photo_data = request.form.get('photo')
    
    if not photo_data:
        return jsonify({'success': False, 'message': '没有收到照片数据'})
    
    # 查找今天的考勤记录
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=current_user.id,
        date=today
    ).first()
    
    if not attendance:
        return jsonify({'success': False, 'message': '今天没有签到记录，无法登记返回'})
    
    # 查找未返回的暂离记录
    active_leave = LeaveRecord.query.filter(
        LeaveRecord.student_id == current_user.id,
        LeaveRecord.attendance_id == attendance.id,
        LeaveRecord.return_time.is_(None)
    ).order_by(LeaveRecord.leave_time.desc()).first()
    
    if not active_leave:
        return jsonify({'success': False, 'message': '没有未完成的暂离记录'})
    
    # 保存照片
    photo_path = save_photo(photo_data, f"return_{current_user.id}")
    
    # 更新返回时间和照片
    now = datetime.now()
    active_leave.return_time = now
    active_leave.return_photo = photo_path
    
    # 计算暂离时长（秒）
    if active_leave.leave_time:
        duration = int((now - active_leave.leave_time).total_seconds())
        active_leave.duration = duration
    
    # 保持向后兼容性
    attendance.return_time = now
    attendance.return_photo = photo_path
    
    db.session.commit()
    
    # 获取暂离开始时间字符串
    leave_time_str = active_leave.leave_time.strftime('%Y-%m-%d %H:%M:%S') if active_leave.leave_time else "未知"
    
    return jsonify({
        'success': True, 
        'message': '返回已记录',
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'leave_time': leave_time_str,
        'duration': active_leave.formatted_duration if active_leave.duration else "未知"
    })

def save_photo(photo_data, prefix):
    """保存照片"""
    if not photo_data:
        return None
    
    # 从Base64数据中提取实际图像数据
    if 'base64,' in photo_data:
        photo_data = photo_data.split('base64,')[1]
    
    import base64
    from datetime import datetime
    from PIL import Image as PILImage
    import io
    
    # 生成文件名
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"{prefix}_{timestamp}.png"
    
    # 确保目录存在
    photos_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'photos')
    if not os.path.exists(photos_dir):
        os.makedirs(photos_dir)
    
    # 将Base64数据解码为图像
    image_data = base64.b64decode(photo_data)
    image = PILImage.open(io.BytesIO(image_data))
    
    # 根据prefix确定水印类型（签到/暂离/返回）和相关学生信息
    watermark_type = "签到"
    student_id = None
    
    if "leave" in prefix:
        watermark_type = "暂离"
        student_id = prefix.split('_')[1] if '_' in prefix else None
    elif "return" in prefix:
        watermark_type = "返回"
        student_id = prefix.split('_')[1] if '_' in prefix else None
    elif "sign_in" in prefix:
        watermark_type = "签到"
        student_id = prefix.split('_')[1] if '_' in prefix else None
    
    # 获取学生信息
    student_name = ""
    class_name = ""
    leave_time_str = ""
    
    # 如果是管理员操作，需要从数据库获取学生信息
    if student_id and isinstance(current_user, ClassroomAdmin):
        student = User.query.filter_by(id=student_id).first()
        if student:
            student_name = student.name
            class_name = student.class_name or ""
            
            # 如果是返回照片，查询暂离时间
            if watermark_type == "返回":
                # 查找最近的未返回暂离记录
                today = datetime.now().date()
                attendance = AttendanceRecord.query.filter_by(
                    student_id=student_id,
                    date=today
                ).first()
                
                if attendance:
                    active_leave = LeaveRecord.query.filter(
                        LeaveRecord.student_id == student_id,
                        LeaveRecord.attendance_id == attendance.id,
                        LeaveRecord.return_time.is_(None)
                    ).order_by(LeaveRecord.leave_time.desc()).first()
                    
                    if active_leave and active_leave.leave_time:
                        leave_time_str = f"暂离时间: {active_leave.leave_time.strftime('%H:%M:%S')}"
    
    # 如果是学生自己操作，直接使用当前用户信息
    elif isinstance(current_user, User) and current_user.role == 'student':
        student_name = current_user.name
        class_name = current_user.class_name or ""
        
        # 如果是返回照片，查询暂离时间
        if watermark_type == "返回":
            # 查找最近的未返回暂离记录
            today = datetime.now().date()
            attendance = AttendanceRecord.query.filter_by(
                student_id=current_user.id,
                date=today
            ).first()
            
            if attendance:
                active_leave = LeaveRecord.query.filter(
                    LeaveRecord.student_id == current_user.id,
                    LeaveRecord.attendance_id == attendance.id,
                    LeaveRecord.return_time.is_(None)
                ).order_by(LeaveRecord.leave_time.desc()).first()
                
                if active_leave and active_leave.leave_time:
                    leave_time_str = f"暂离时间: {active_leave.leave_time.strftime('%H:%M:%S')}"
    
    # 生成时间文本并添加水印
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 组合水印文本，根据不同类型生成不同内容
    timestamp_text = f"{class_name} {student_name}\n{watermark_type}: {current_time}"
    if leave_time_str and watermark_type == "返回":
        timestamp_text = f"{class_name} {student_name}\n{leave_time_str}\n{watermark_type}: {current_time}"
    
    img_with_watermark = add_timestamp_watermark(image, timestamp_text, position='bottom', no_background=True)
    
    # 保存文件，使用高质量设置
    filepath = os.path.join(photos_dir, filename)
    
    # 转换为RGB格式（处理可能的RGBA格式）并以高质量保存
    if img_with_watermark.mode == 'RGBA':
        # 创建白色背景
        background = PILImage.new('RGB', img_with_watermark.size, (255, 255, 255))
        # 将RGBA图像粘贴到白色背景上
        background.paste(img_with_watermark, mask=img_with_watermark.split()[3])
        background.save(filepath, "PNG", quality=100, optimize=True)
    else:
        img_with_watermark.save(filepath, "PNG", quality=100, optimize=True)
    
    # 返回相对路径
    return os.path.join('photos', filename)

@app.route('/students')
@login_required
def student_list():
    """学生名单管理页面"""
    if current_user.role != 'teacher':
        flash('无权访问学生名单管理', 'danger')
        return redirect(url_for('index'))
    
    # 获取所有学生
    students = User.query.filter_by(role='student').order_by(User.class_name, User.username).all()
    
    # 获取所有班级
    classes = db.session.query(User.class_name).filter(
        User.role == 'student',
        User.class_name.isnot(None)
    ).distinct().all()
    classes = [c[0] for c in classes if c[0]]
    
    # 导入表单
    import_form = ImportStudentsForm()
    
    return render_template('student_list.html', 
                         students=students, 
                         classes=classes,
                         import_form=import_form)

@app.route('/download_template')
@login_required
def download_template():
    """下载Excel模板"""
    if current_user.role != 'teacher':
        flash('无权下载模板', 'danger')
        return redirect(url_for('index'))
    
    wb = generate_template()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='学生名单导入模板.xlsx'
    )

@app.route('/import_students', methods=['POST'])
@login_required
def import_students():
    """导入学生名单"""
    if current_user.role != 'teacher':
        flash('无权导入学生名单', 'danger')
        return redirect(url_for('index'))
    
    form = ImportStudentsForm()
    if form.validate_on_submit():
        file = form.file.data
        filename = secure_filename(file.filename)
        
        # 确保上传目录存在
        uploads_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'excel')
        if not os.path.exists(uploads_dir):
            os.makedirs(uploads_dir)
        
        # 保存上传的文件
        filepath = os.path.join(uploads_dir, filename)
        file.save(filepath)
        
        # 导入学生数据
        success, message = import_students_from_excel(filepath)
        
        if success:
            flash(message, 'success')
        else:
            flash(message, 'danger')
    
    return redirect(url_for('student_list'))

@app.route('/attendance')
@login_required
def attendance_records():
    """考勤记录页面"""
    if current_user.role != 'teacher':
        flash('无权访问考勤记录', 'danger')
        return redirect(url_for('index'))
    
    # 获取今天的日期
    today = datetime.now().date()
    # 默认查询今天的考勤记录
    date = request.args.get('date', today.strftime('%Y-%m-%d'))
    try:
        query_date = datetime.strptime(date, '%Y-%m-%d').date()
    except:
        query_date = today
    
    # 获取班级过滤
    class_name = request.args.get('class', '')
    
    # 查询考勤记录
    query = db.session.query(
        User, AttendanceRecord
    ).outerjoin(
        AttendanceRecord, 
        db.and_(
            User.id == AttendanceRecord.student_id,
            AttendanceRecord.date == query_date
        )
    ).filter(
        User.role == 'student'
    )
    
    # 如果指定了班级，则过滤
    if class_name:
        query = query.filter(User.class_name == class_name)
    
    # 执行查询
    records = query.all()
    
    # 获取所有班级
    classes = db.session.query(User.class_name).filter(
        User.role == 'student',
        User.class_name.isnot(None)
    ).distinct().all()
    classes = [c[0] for c in classes if c[0]]
    
    # 导出表单
    export_form = ExportAttendanceForm()
    export_form.class_name.choices = [('', '全部班级')] + [(c, c) for c in classes]
    
    # 默认设置为最近一周
    if not export_form.start_date.data:
        export_form.start_date.data = today - timedelta(days=6)
    if not export_form.end_date.data:
        export_form.end_date.data = today
    
    return render_template('attendance_records.html',
                         records=records,
                         classes=classes,
                         current_class=class_name,
                         date=query_date,
                         export_form=export_form)

@app.route('/export_attendance', methods=['POST'])
@login_required
def export_attendance():
    """导出考勤记录"""
    if current_user.role != 'teacher':
        flash('无权导出考勤记录', 'danger')
        return redirect(url_for('index'))
    
    form = ExportAttendanceForm()
    # 获取所有班级
    classes = db.session.query(User.class_name).filter(
        User.role == 'student',
        User.class_name.isnot(None)
    ).distinct().all()
    classes = [c[0] for c in classes if c[0]]
    form.class_name.choices = [('', '全部班级')] + [(c, c) for c in classes]
    
    if form.validate_on_submit():
        start_date = form.start_date.data
        end_date = form.end_date.data
        class_name = form.class_name.data if form.class_name.data else None
        export_photos = form.export_photos.data
        
        # 导出考勤记录
        filename = export_attendance_records(start_date, end_date, class_name, export_photos)
        
        if not filename:
            flash('没有找到符合条件的考勤记录，或导出过程中出现错误', 'warning')
            return redirect(url_for('attendance_records'))
        
        # 如果导出照片，也提示用户照片已导出
        if export_photos:
            photos_dir_name = filename.replace('.xlsx', '_照片')
            flash(f'照片已嵌入Excel文件，完整照片集可在 uploads/exports/{photos_dir_name} 文件夹中查看', 'success')
        
        # 构建文件路径并发送文件
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], 'exports', filename)
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    # 如果表单验证失败，显示错误
    for field, errors in form.errors.items():
        for error in errors:
            flash(f'{form[field].label.text}: {error}', 'danger')
    
    return redirect(url_for('attendance_records'))

@app.route('/reset_db')
def reset_db():
    """重建数据库"""
    with app.app_context():
        db.drop_all()  # 删除所有表
        db.create_all()  # 重新创建所有表
        
        # 创建管理员账号
        admin = User(
            username='admin',
            name='管理员',
            role='teacher'
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
            
        return '数据库已重置，并创建了管理员账号'

# 在应用启动前初始化数据库
with app.app_context():
    try:
        # 尝试查询管理员账号，如果因为表结构变化而出错，则重建表
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                name='管理员',
                role='teacher'
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
        print("数据库已自动初始化")
    except Exception as e:
        # 如果出错，可能是表结构变化，尝试重建表
        print("数据库查询出错，尝试重建数据库:", str(e))
        db.drop_all()
        db.create_all()
        # 创建管理员账号
        admin = User(
            username='admin',
            name='管理员',
            role='teacher'
        )
        admin.set_password('admin123')
        db.session.add(admin)
        db.session.commit()
        print("数据库结构已更新，已重新创建管理员账号")

def admin_required(f):
    """系统管理员权限检查装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'teacher':
            flash('需要系统管理员权限', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/admin/accounts')
@admin_required
def admin_accounts():
    """教室管理员账号管理页面"""
    admin_accounts = ClassroomAdmin.query.all()
    return render_template('admin_accounts.html', admin_accounts=admin_accounts)

@app.route('/api/admin/accounts', methods=['POST'])
@admin_required
def create_admin_account():
    """创建教室管理员账号"""
    data = request.get_json()
    
    # 检查必要字段
    if not data.get('classroom_name') or not data.get('username'):
        return jsonify({'error': '教室名称和账号不能为空'}), 400
        
    # 规范化用户名和教室名称
    username = data['username'].strip()  # 移除lower()，保留原始大小写
    classroom_name = data['classroom_name'].upper().strip()
        
    # 检查账号是否已存在（不区分大小写进行检查）
    if ClassroomAdmin.query.filter(db.func.lower(ClassroomAdmin.username) == username.lower()).first():
        return jsonify({'error': '该账号已存在'}), 400
        
    # 检查教室名称是否已存在
    existing_classroom = ClassroomAdmin.query.filter_by(classroom_name=classroom_name).first()
    if existing_classroom:
        return jsonify({'error': f'教室 {classroom_name} 已存在，管理员为：{existing_classroom.username}'}), 400
        
    # 创建新账号
    admin = ClassroomAdmin(
        username=username,
        classroom_name=classroom_name,
    )
    
    # 设置密码（如果未提供则使用默认密码）
    password = data.get('password') or '123456'
    admin.set_password(password)
    
    try:
        db.session.add(admin)
        db.session.commit()
        
        # 获取管理员ID，用于后续关联学生
        admin_id = admin.id
        
        # 关联已有该教室位置的学生（根据教室名称匹配）
        students_updated = 0
        
        # 查找所有教室位置与新创建的教室名称匹配的学生
        matching_students = User.query.filter(
            User.role == 'student',
            User.classroom_location.isnot(None),
            # 使用函数将数据库字段转为大写再比较，或者直接使用教室名称的大写形式
            db.func.upper(User.classroom_location) == classroom_name
        ).all()
        
        logger.info(f"找到 {len(matching_students)} 名学生的教室位置与 {classroom_name} 匹配")
        
        for student in matching_students:
            student.classroom_id = admin_id
            # 确保教室位置使用标准格式
            if student.classroom_location != classroom_name:
                student.classroom_location = classroom_name
            students_updated += 1
            logger.info(f"关联学生 {student.name} 到教室 {classroom_name}，管理员ID: {admin_id}")
            
        if students_updated > 0:
            db.session.commit()
            logger.info(f"已将 {students_updated} 名学生关联到新创建的教室 {classroom_name}")
        
        # 在返回结果中包含关联信息
        result = admin.to_dict()
        result['students_linked'] = students_updated
        
        return jsonify(result), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': '创建失败：' + str(e)}), 500

@app.route('/api/admin/accounts/<int:admin_id>', methods=['PUT'])
@admin_required
def update_admin_account(admin_id):
    """更新教室管理员账号信息"""
    try:
        logger.info(f"正在更新教室管理员账号 ID: {admin_id}")
        
        # 检查管理员是否存在
        admin = ClassroomAdmin.query.get(admin_id)
        if not admin:
            logger.error(f"管理员ID {admin_id} 不存在")
            return jsonify({'error': f'管理员ID {admin_id} 不存在'}), 404
        
        # 获取请求数据
        data = request.get_json()
        if not data:
            logger.error("请求数据为空或格式错误")
            return jsonify({'error': '请求数据为空或格式错误'}), 400
            
        logger.info(f"收到更新请求数据: {data}")
        
        # 记录原始教室名称和用户名，用于后续更新学生关联
        old_classroom_name = admin.classroom_name
        old_username = admin.username
        logger.info(f"原有教室名称: {old_classroom_name}, 原有用户名: {old_username}")
        
        # 更新教室名称
        new_classroom_name = None
        if 'classroom_name' in data and data['classroom_name'].strip():
            new_classroom_name = data['classroom_name'].upper().strip()
            logger.info(f"准备更新教室名称: {old_classroom_name} -> {new_classroom_name}")
            
            # 检查名称是否已被其他管理员使用
            if new_classroom_name != old_classroom_name:
                existing = ClassroomAdmin.query.filter(
                    ClassroomAdmin.classroom_name == new_classroom_name,
                    ClassroomAdmin.id != admin_id
                ).first()
                
                if existing:
                    logger.error(f"教室名称 {new_classroom_name} 已被管理员 {existing.username} 使用")
                    return jsonify({'error': f'教室名称 {new_classroom_name} 已被其他管理员使用'}), 400
                
                # 更新教室名称
                admin.classroom_name = new_classroom_name
                logger.info(f"更新教室名称: {old_classroom_name} -> {new_classroom_name}")
                
                # 更新关联到此教室的学生记录
                students_to_update = User.query.filter_by(classroom_id=admin_id).all()
                logger.info(f"找到 {len(students_to_update)} 名关联学生需要更新教室位置")
                
                for student in students_to_update:
                    student.classroom_location = new_classroom_name
                    logger.info(f"更新学生 {student.name} 的教室位置: {old_classroom_name} -> {new_classroom_name}")
        else:
            logger.info("未提供教室名称或名称为空，跳过更新")
        
        # 更新账号名称
        new_username = None
        if 'username' in data and data['username'].strip():
            new_username = data['username'].strip()  # 移除lower()，保留原始大小写
            logger.info(f"准备更新账号: {old_username} -> {new_username}")
            
            # 如果用户名有变化，检查是否已存在
            if new_username.lower() != old_username.lower():
                existing = ClassroomAdmin.query.filter(
                    db.func.lower(ClassroomAdmin.username) == new_username.lower(),
                    ClassroomAdmin.id != admin_id
                ).first()
                
                if existing:
                    logger.error(f"账号 {new_username} 已存在")
                    return jsonify({'error': f'账号 {new_username} 已存在'}), 400
                
                admin.username = new_username
                logger.info(f"更新账号: {old_username} -> {new_username}")
        else:
            logger.info("未提供账号或账号为空，跳过更新")
        
        # 确认是否有任何字段被更新
        if new_classroom_name is None and new_username is None:
            logger.warning("没有任何字段需要更新")
            return jsonify({'warning': '没有提供需要更新的信息'}), 200
            
        # 提交更改
        db.session.commit()
        logger.info(f"成功更新教室管理员 ID: {admin_id}")
        
        # 返回更新后的数据
        result = admin.to_dict()
        logger.info(f"返回数据: {result}")
        return jsonify(result)
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        logger.exception(f"更新教室管理员失败: {error_msg}")
        return jsonify({'error': f'更新失败: {error_msg}'}), 500

@app.route('/api/admin/accounts/<int:admin_id>/toggle-status', methods=['POST'])
@admin_required
def toggle_admin_status(admin_id):
    """启用/禁用教室管理员账号"""
    admin = ClassroomAdmin.query.get_or_404(admin_id)
    data = request.get_json()
    
    admin.is_active = data.get('is_active', not admin.is_active)
    
    try:
        db.session.commit()
        return jsonify(admin.to_dict())
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': '操作失败：' + str(e)}), 500

@app.route('/api/admin/accounts/reset-password', methods=['POST'])
@admin_required
def reset_admin_password():
    """重置教室管理员密码"""
    data = request.get_json()
    admin_id = data.get('admin_id')
    new_password = data.get('new_password', '123456')  # 如果未提供新密码，则使用默认密码
    
    admin = ClassroomAdmin.query.get_or_404(admin_id)
    admin.set_password(new_password)
    
    try:
        db.session.commit()
        return jsonify({'message': '密码重置成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': '密码重置失败：' + str(e)}), 500

@app.route('/admin/accounts/template')
@admin_required
def download_admin_template():
    """下载教室管理员账号导入模板"""
    wb = generate_admin_template()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='教室管理员账号导入模板.xlsx'
    )

@app.route('/admin/accounts/import', methods=['POST'])
@admin_required
def import_admin_accounts_route():
    """导入教室管理员账号"""
    if 'file' not in request.files:
        flash('请选择要导入的文件', 'danger')
        return redirect(url_for('admin_accounts'))
    
    file = request.files['file']
    if file.filename == '':
        flash('未选择文件', 'danger')
        return redirect(url_for('admin_accounts'))
    
    if not file.filename.endswith('.xlsx'):
        flash('请上传Excel文件(.xlsx格式)', 'danger')
        return redirect(url_for('admin_accounts'))
    
    # 保存上传的文件
    uploads_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'excel')
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)
    
    filepath = os.path.join(uploads_dir, secure_filename(file.filename))
    file.save(filepath)
    
    # 导入账号
    success, message = import_admin_accounts(filepath)
    
    if success:
        flash(message, 'success')
    else:
        flash(message, 'danger')
    
    return redirect(url_for('admin_accounts'))

@app.route('/admin/accounts/export')
@admin_required
def export_admin_accounts_route():
    """导出教室管理员账号"""
    wb = export_admin_accounts()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'教室管理员账号_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/classroom_admin')
@login_required
def classroom_admin_dashboard():
    """教室管理员仪表盘页面"""
    if not isinstance(current_user, ClassroomAdmin):
        flash('无权访问教室管理员页面', 'danger')
        return redirect(url_for('index'))
    
    # 获取今天的日期
    today = datetime.now().date()
    
    # 直接使用教室名称匹配学生，而不是使用classroom_id关联
    # 获取该教室的学生列表（通过教室位置与管理员用户名匹配）
    students = User.query.filter(
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).all()
    
    # 获取已签到学生数
    signed_in = db.session.query(User).join(
        AttendanceRecord, User.id == AttendanceRecord.student_id
    ).filter(
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username),
        AttendanceRecord.date == today,
        AttendanceRecord.sign_in_time.isnot(None)
    ).count()
    
    # 计算签到率
    total_students = len(students)
    signed_in_rate = round(signed_in / total_students * 100, 2) if total_students > 0 else 0
    
    return render_template('classroom_admin_dashboard.html', 
                         classroom_name=current_user.classroom_name,
                         total_students=total_students,
                         signed_in=signed_in,
                         signed_in_rate=signed_in_rate,
                         students=students,
                         today=today)

# API 路由
@app.route('/api/sign_in', methods=['POST'])
@login_required
def api_sign_in():
    """教室管理员API：学生签到"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 同时支持JSON和表单数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    # 尝试从表单中获取数据
    student_id = data.get('student_id') or request.form.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
    
    # 查询该学生是否属于该教室管理员（通过教室位置与管理员用户名匹配）
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 记录签到信息
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance:
        # 如果没有签到记录，创建一个
        attendance = AttendanceRecord(
            student_id=student_id,
            student_name=student.name,
            class_name=student.class_name,
            date=today
        )
        db.session.add(attendance)
    
    attendance.sign_in_time = datetime.now()
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': '签到成功',
        'student_name': student.name,
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

@app.route('/api/leave', methods=['POST'])
@login_required
def api_leave():
    """教室管理员API：学生暂离"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 同时支持JSON和表单数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    # 尝试从表单中获取数据
    student_id = data.get('student_id') or request.form.get('student_id')
    reason = data.get('reason') or request.form.get('reason', '')  # 可选的暂离原因
    photo = data.get('photo') or request.form.get('photo')  # 增加对照片的支持
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
    
    # 查询该学生是否属于该教室管理员（通过教室位置与管理员用户名匹配）
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 记录暂离信息
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance or not attendance.sign_in_time:
        return jsonify({'success': False, 'message': '该学生今天尚未签到，无法暂离'})
    
    # 检查是否有未返回的暂离记录
    active_leave = LeaveRecord.query.filter(
        LeaveRecord.student_id == student_id,
        LeaveRecord.attendance_id == attendance.id,
        LeaveRecord.return_time.is_(None)
    ).first()
    
    if active_leave:
        return jsonify({'success': False, 'message': '该学生有未完成的暂离记录，请先登记返回'})
    
    # 如果提供了照片，保存照片
    leave_photo = None
    if photo:
        leave_photo = save_photo(photo, f"leave_{student_id}")
    
    # 创建新的暂离记录
    leave_record = LeaveRecord(
        student_id=student_id,
        student_name=student.name,
        class_name=student.class_name,
        attendance_id=attendance.id,
        leave_time=datetime.now(),
        reason=reason,
        leave_photo=leave_photo
    )
    
    db.session.add(leave_record)
    
    # 保持向后兼容性
    attendance.leave_time = datetime.now()
    if leave_photo:
        attendance.leave_photo = leave_photo
    
    try:
        db.session.commit()
        
        # 记录到日志
        logger.info(f"学生 {student.name}(ID:{student_id}) 暂离登记成功，由 {current_user.username} 操作")
        
        return jsonify({
            'success': True, 
            'message': '暂离登记成功',
            'student_name': student.name,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'leave_id': leave_record.id
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"暂离登记失败: {str(e)}")
        return jsonify({'success': False, 'message': f'暂离登记失败: {str(e)}'})

@app.route('/api/return', methods=['POST'])
@login_required
def api_return():
    """教室管理员API：学生暂离归来"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 同时支持JSON和表单数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    # 尝试从表单中获取数据
    student_id = data.get('student_id') or request.form.get('student_id')
    photo = data.get('photo') or request.form.get('photo')
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
    
    if not photo:
        return jsonify({'success': False, 'message': '未提供照片数据'})
    
    # 查询该学生是否属于该教室管理员（通过教室位置与管理员用户名匹配）
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 查找最近的活跃暂离记录
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance:
        return jsonify({'success': False, 'message': '该学生今天没有考勤记录'})
    
    # 查找未返回的暂离记录
    active_leave = LeaveRecord.query.filter(
        LeaveRecord.student_id == student_id,
        LeaveRecord.attendance_id == attendance.id,
        LeaveRecord.return_time.is_(None)
    ).order_by(LeaveRecord.leave_time.desc()).first()
    
    if not active_leave:
        return jsonify({'success': False, 'message': '该学生没有未完成的暂离记录'})
    
    # 保存照片
    photo_path = save_photo(photo, f"return_{student_id}")
    
    # 更新返回时间和照片
    now = datetime.now()
    active_leave.return_time = now
    active_leave.return_photo = photo_path
    
    # 计算暂离时长（秒）
    if active_leave.leave_time:
        duration = int((now - active_leave.leave_time).total_seconds())
        active_leave.duration = duration
    
    # 保持向后兼容性 - 这里只更新记录，但不影响状态判断
    # 注意：暂离状态现在由最新的LeaveRecord决定，而不是由AttendanceRecord决定
    if active_leave.is_active:
        attendance.return_time = now
        attendance.return_photo = photo_path
    
    try:
        db.session.commit()
        
        # 记录到日志
        duration_str = active_leave.formatted_duration if active_leave.duration else "未知"
        leave_time_str = active_leave.leave_time.strftime('%H:%M:%S') if active_leave.leave_time else "未知"
        logger.info(f"学生 {student.name}(ID:{student_id}) 暂离归来登记成功，暂离开始: {leave_time_str}，暂离时长: {duration_str}，由 {current_user.username} 操作")
        
        return jsonify({
            'success': True, 
            'message': '归来登记成功',
            'student_name': student.name,
            'leave_time': leave_time_str,
            'time': now.strftime('%Y-%m-%d %H:%M:%S'),
            'duration': active_leave.formatted_duration if active_leave.duration else "未知"
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"归来登记失败: {str(e)}")
        return jsonify({'success': False, 'message': f'归来登记失败: {str(e)}'})

@app.route('/api/absence', methods=['POST'])
@login_required
def api_absence():
    """教室管理员API：学生请假"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 同时支持JSON和表单数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    # 尝试从表单中获取数据
    student_id = data.get('student_id') or request.form.get('student_id')
    absence_type = data.get('absence_type') or request.form.get('absence_type')
    reason = data.get('reason') or request.form.get('reason', '')
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
        
    if not absence_type:
        return jsonify({'success': False, 'message': '未提供请假类型'})
    
    # 查询该学生是否属于该教室管理员（通过教室位置与管理员用户名匹配）
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 检查今天是否已有请假记录
    today = datetime.now().date()
    existing_absence = AbsenceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if existing_absence:
        return jsonify({'success': False, 'message': f'该学生今天已请假({existing_absence.absence_type})'})
    
    # 创建新的请假记录
    absence = AbsenceRecord(
        student_id=student_id,
        student_name=student.name,
        class_name=student.class_name,
        date=today,
        absence_type=absence_type,
        reason=reason,
        approved_by=current_user.username
    )
    
    try:
        db.session.add(absence)
        db.session.commit()
        
        # 记录到日志
        logger.info(f"学生 {student.name}(ID:{student_id}) 请假登记成功，类型: {absence_type}，原因: {reason}，由 {current_user.username} 操作")
        
        return jsonify({
            'success': True, 
            'message': '请假登记成功',
            'student_name': student.name,
            'absence_type': absence_type,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"请假登记失败: {str(e)}")
        return jsonify({'success': False, 'message': f'请假登记失败: {str(e)}'})

@app.route('/api/absence/cancel', methods=['POST'])
@login_required
def api_cancel_absence():
    """教室管理员API：取消学生请假"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 获取数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    student_id = data.get('student_id') or request.form.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
    
    # 查询该学生是否属于该教室管理员
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 查找今天的请假记录
    today = datetime.now().date()
    absence = AbsenceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not absence:
        return jsonify({'success': False, 'message': '该学生今天没有请假记录'})
    
    # 记录请假信息以便日志记录
    absence_info = {
        'student_name': absence.student_name,
        'absence_type': absence.absence_type,
        'reason': absence.reason
    }
    
    try:
        # 删除请假记录
        db.session.delete(absence)
        db.session.commit()
        
        # 记录到日志
        logger.info(f"学生 {absence_info['student_name']}(ID:{student_id}) 请假取消成功，原类型: {absence_info['absence_type']}，由 {current_user.username} 操作")
        
        return jsonify({
            'success': True,
            'message': '请假记录已取消',
            'student_name': absence_info['student_name']
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"取消请假失败: {str(e)}")
        return jsonify({'success': False, 'message': f'取消请假失败: {str(e)}'})

@app.route('/api/sign_out', methods=['POST'])
@login_required
def api_sign_out():
    """教室管理员API：学生签出"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 获取数据
    try:
        data = request.json or {}
    except:
        data = {}
    
    student_id = data.get('student_id') or request.form.get('student_id')
    
    if not student_id:
        return jsonify({'success': False, 'message': '未提供学生ID'})
    
    # 查询该学生是否属于该教室管理员
    student = User.query.filter(
        User.id == student_id,
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).first()
    
    if not student:
        return jsonify({'success': False, 'message': '该学生不属于您管理的教室'})
    
    # 查找今天的签到记录
    today = datetime.now().date()
    attendance = AttendanceRecord.query.filter_by(
        student_id=student_id,
        date=today
    ).first()
    
    if not attendance:
        return jsonify({'success': False, 'message': '该学生今天没有签到记录'})
    
    if not attendance.sign_in_time:
        return jsonify({'success': False, 'message': '该学生今天尚未签到，无法签出'})
    
    if attendance.sign_out_time:
        return jsonify({'success': False, 'message': '该学生今天已经签出'})
    
    # 检查是否有未完成的暂离记录
    active_leave = LeaveRecord.query.filter(
        LeaveRecord.student_id == student_id,
        LeaveRecord.attendance_id == attendance.id,
        LeaveRecord.return_time.is_(None)
    ).first()
    
    if active_leave:
        return jsonify({'success': False, 'message': '该学生有未完成的暂离记录，请先登记返回'})
    
    try:
        # 更新签出时间
        attendance.sign_out_time = datetime.now()
        db.session.commit()
        
        # 记录到日志
        sign_in_time_str = attendance.sign_in_time.strftime('%H:%M:%S') if attendance.sign_in_time else "未知"
        sign_out_time_str = attendance.sign_out_time.strftime('%H:%M:%S') if attendance.sign_out_time else "未知"
        logger.info(f"学生 {student.name}(ID:{student_id}) 签出成功，签到时间: {sign_in_time_str}，签出时间: {sign_out_time_str}，由 {current_user.username} 操作")
        
        return jsonify({
            'success': True,
            'message': '签出成功',
            'student_name': student.name,
            'sign_in_time': sign_in_time_str,
            'sign_out_time': sign_out_time_str
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"签出失败: {str(e)}")
        return jsonify({'success': False, 'message': f'签出失败: {str(e)}'})

@app.route('/api/sign_out_all', methods=['POST'])
@login_required
def api_sign_out_all():
    """教室管理员API：所有学生一键签出"""
    if not isinstance(current_user, ClassroomAdmin):
        return jsonify({'success': False, 'message': '只有教室管理员可以使用此功能'})
    
    # 查找今天已签到但未签出的学生
    today = datetime.now().date()
    
    # 获取该教室的所有学生的ID
    students = User.query.filter(
        User.role == 'student',
        db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
    ).all()
    student_ids = [s.id for s in students]
    
    # 查找这些学生今天已签到但未签出的记录
    attendances = AttendanceRecord.query.filter(
        AttendanceRecord.student_id.in_(student_ids),
        AttendanceRecord.date == today,
        AttendanceRecord.sign_in_time.isnot(None),
        AttendanceRecord.sign_out_time.is_(None)
    ).all()
    
    if not attendances:
        return jsonify({'success': False, 'message': '没有找到需要签出的学生记录'})
    
    # 检查是否有未完成的暂离记录
    student_with_active_leave = []
    for attendance in attendances:
        active_leave = LeaveRecord.query.filter(
            LeaveRecord.student_id == attendance.student_id,
            LeaveRecord.attendance_id == attendance.id,
            LeaveRecord.return_time.is_(None)
        ).first()
        
        if active_leave:
            student_with_active_leave.append(attendance.student_name)
    
    if student_with_active_leave:
        return jsonify({
            'success': False, 
            'message': f'以下学生有未完成的暂离记录，请先处理：{", ".join(student_with_active_leave)}'
        })
    
    try:
        # 签出所有学生
        now = datetime.now()
        signed_out_count = 0
        
        for attendance in attendances:
            attendance.sign_out_time = now
            signed_out_count += 1
            
            # 记录到日志
            logger.info(f"学生 {attendance.student_name}(ID:{attendance.student_id}) 被一键签出，操作者: {current_user.username}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'成功签出 {signed_out_count} 名学生',
            'count': signed_out_count,
            'time': now.strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"一键签出失败: {str(e)}")
        return jsonify({'success': False, 'message': f'一键签出失败: {str(e)}'})

@app.route('/api/absences', methods=['GET'])
@login_required
def api_get_absences():
    """获取请假记录"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': '请先登录'})
    
    # 获取请假日期，默认为今天
    date_str = request.args.get('date')
    try:
        if date_str:
            query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            query_date = datetime.now().date()
    except ValueError:
        return jsonify({'success': False, 'message': '日期格式错误'})
    
    try:
        # 查询条件
        query = AbsenceRecord.query.filter(AbsenceRecord.date == query_date)
        
        # 如果是教室管理员，只显示该教室的学生
        if isinstance(current_user, ClassroomAdmin):
            # 获取该教室的所有学生
            student_ids = db.session.query(User.id).filter(
                User.role == 'student',
                db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
            ).all()
            student_ids = [s[0] for s in student_ids]
            query = query.filter(AbsenceRecord.student_id.in_(student_ids))
        
        # 如果是教师，可以查看所有人
        absences = query.order_by(AbsenceRecord.class_name, AbsenceRecord.student_name).all()
        
        # 转换为字典列表
        result = []
        for absence in absences:
            # 确保学生对象已关联
            if not absence.student:
                student = User.query.get(absence.student_id)
                absence.student = student
            
            absence_dict = absence.to_dict()
            
            # 添加用户信息
            if absence.student:
                absence_dict['username'] = absence.student.username
            
            result.append(absence_dict)
        
        return jsonify({
            'success': True,
            'absences': result,
            'date': query_date.strftime('%Y-%m-%d'),
            'count': len(result)
        })
    except Exception as e:
        logger.error(f"获取请假记录失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取请假记录失败: {str(e)}'})

@app.route('/api/export_absences', methods=['GET'])
@login_required
def api_export_absences():
    """导出请假记录"""
    if not current_user.is_authenticated:
        return jsonify({'success': False, 'message': '请先登录'})
    
    # 获取请假日期，默认为今天
    date_str = request.args.get('date')
    try:
        if date_str:
            query_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        else:
            query_date = datetime.now().date()
    except ValueError:
        return jsonify({'success': False, 'message': '日期格式错误'})
    
    try:
        # 查询条件
        query = AbsenceRecord.query.filter(AbsenceRecord.date == query_date)
        
        # 获取班级参数
        class_name = request.args.get('class')
        if class_name:
            query = query.filter(AbsenceRecord.class_name == class_name)
        
        # 如果是教室管理员，只显示该教室的学生
        if isinstance(current_user, ClassroomAdmin):
            # 获取该教室的所有学生
            student_ids = db.session.query(User.id).filter(
                User.role == 'student',
                db.func.upper(User.classroom_location) == db.func.upper(current_user.username)
            ).all()
            student_ids = [s[0] for s in student_ids]
            query = query.filter(AbsenceRecord.student_id.in_(student_ids))
        
        # 执行查询
        absences = query.order_by(AbsenceRecord.class_name, AbsenceRecord.student_name).all()
        
        if not absences:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': '没有找到符合条件的请假记录'})
            else:
                flash('没有找到符合条件的请假记录', 'warning')
                return redirect(url_for('attendance_records'))
        
        # 创建Excel工作簿
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"请假记录_{query_date.strftime('%Y%m%d')}"
        
        # 设置标题和样式
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加标题行
        headers = ['序号', '班级', '学号', '姓名', '请假类型', '请假原因', '批准人', '登记时间']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加数据
        for row, absence in enumerate(absences, start=2):
            # 查询学生信息
            student = User.query.get(absence.student_id)
            username = student.username if student else ""
            
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=absence.class_name or "")
            ws.cell(row=row, column=3, value=username)
            ws.cell(row=row, column=4, value=absence.student_name)
            ws.cell(row=row, column=5, value=absence.absence_type)
            ws.cell(row=row, column=6, value=absence.reason or "")
            ws.cell(row=row, column=7, value=absence.approved_by or "")
            ws.cell(row=row, column=8, value=absence.created_at.strftime('%Y-%m-%d %H:%M:%S') if absence.created_at else "")
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 设置每一列的对齐方式
        for row in ws.iter_rows(min_row=2, max_row=len(absences) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 创建文件名
        filename = f"请假记录_{query_date.strftime('%Y%m%d')}.xlsx"
        if class_name:
            filename = f"请假记录_{class_name}_{query_date.strftime('%Y%m%d')}.xlsx"
        
        # 确保目录存在
        exports_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'exports')
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
        
        # 保存文件
        filepath = os.path.join(exports_dir, filename)
        wb.save(filepath)
        
        # 发送文件
        return send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"导出请假记录失败: {str(e)}")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})
        else:
            flash(f'导出请假记录失败: {str(e)}', 'danger')
            return redirect(url_for('attendance_records'))

@app.route('/api/students', methods=['POST'])
@admin_required
def add_student():
    """添加单个学生"""
    try:
        # 获取表单数据
        class_name = request.form.get('class_name')
        username = request.form.get('username')
        name = request.form.get('name')
        password = request.form.get('password')
        classroom_location = request.form.get('classroom_location')

        # 验证必填字段
        if not all([class_name, username, name, password]):
            return jsonify({'success': False, 'message': '请填写所有必填字段'})

        # 检查学号是否已存在
        if User.query.filter_by(username=username).first():
            return jsonify({'success': False, 'message': '该学号已存在'})

        # 创建新学生
        student = User(
            username=username,
            name=name,
            class_name=class_name,
            role='student'
        )

        # 如果提供了教室位置，规范化并保存
        if classroom_location:
            # 统一转为大写进行保存
            normalized_location = classroom_location.strip().upper()
            student.classroom_location = normalized_location  # 确保存储规范化的位置
            
            # 查找匹配的教室管理员（仅用于日志记录）
            admin = ClassroomAdmin.query.filter_by(classroom_name=normalized_location).first()
            
            if admin:
                logger.info(f"学生 {name}(ID:{username}) 的教室位置 {normalized_location} 匹配教室管理员: {admin.username}")
            else:
                logger.warning(f"未找到教室名称为 {normalized_location} 的管理员，学生 {name} 将无法自动关联教室")
        else:
            logger.info(f"学生 {name} 未提供教室位置")

        # 设置密码
        #student.set_password(password)

        # 保存到数据库
        db.session.add(student)
        db.session.commit()

        # 返回结果包含更详细信息，便于调试
        result_message = '学生添加成功'
        if classroom_location:
            result_message += f'，教室位置：{student.classroom_location}'

        return jsonify({'success': True, 'message': result_message})
    except Exception as e:
        db.session.rollback()
        logger.error(f"添加学生失败: {str(e)}")
        return jsonify({'success': False, 'message': f'添加失败: {str(e)}'})

@app.route('/api/students/<int:student_id>', methods=['DELETE'])
@admin_required
def delete_student(student_id):
    """删除单个学生"""
    try:
        # 查找学生
        student = User.query.get_or_404(student_id)
        
        # 检查是否是学生账号
        if student.role != 'student':
            return jsonify({'success': False, 'message': '只能删除学生账号'})
        
        # 删除相关的考勤记录
        AttendanceRecord.query.filter_by(student_id=student_id).delete()
        
        # 删除学生
        db.session.delete(student)
        db.session.commit()
        
        return jsonify({'success': True, 'message': '学生删除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"删除学生失败: {str(e)}")
        return jsonify({'success': False, 'message': f'删除失败: {str(e)}'})

@app.route('/api/sync_student_classrooms', methods=['POST'])
@admin_required
def sync_student_classrooms():
    """同步所有学生与教室管理员的关联关系"""
    try:
        # 获取所有教室管理员，建立教室名称到ID的映射
        # 注意：这里使用教室名称(classroom_name)作为键，而不是账号(username)
        classroom_admins = {}
        for admin in ClassroomAdmin.query.all():
            classroom_admins[admin.classroom_name] = admin.id
            logger.info(f"教室: {admin.classroom_name}, 管理员ID: {admin.id}, 账号: {admin.username}")
        
        # 获取所有有教室位置但没有关联教室管理员的学生
        students = User.query.filter(
            User.role == 'student',
            User.classroom_location.isnot(None),
            (User.classroom_id.is_(None) | (User.classroom_id == 0))
        ).all()
        
        updated_count = 0
        for student in students:
            # 规范化教室位置以匹配教室名称格式
            normalized_location = student.classroom_location.strip().upper()
            # 更新规范化的教室位置
            if student.classroom_location != normalized_location:
                student.classroom_location = normalized_location
            
            # 如果找到匹配的教室管理员，建立关联
            if normalized_location in classroom_admins:
                admin_id = classroom_admins[normalized_location]
                student.classroom_id = admin_id
                updated_count += 1
                logger.info(f"关联学生 {student.name} 到教室 {normalized_location}，管理员ID: {admin_id}")
            else:
                logger.warning(f"找不到匹配的教室: 学生 {student.name} 的教室位置 {normalized_location} 没有对应的教室管理员")
        
        # 提交更改
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'成功同步 {updated_count} 名学生的教室关联',
            'total_processed': len(students),
            'updated': updated_count,
            'available_classrooms': list(classroom_admins.keys())
        })
    except Exception as e:
        db.session.rollback()
        logger.error(f"同步学生教室关联失败: {str(e)}")
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'})

@app.route('/api/admin/accounts/<int:admin_id>', methods=['DELETE'])
@admin_required
def delete_admin_account(admin_id):
    """删除教室管理员账号"""
    try:
        logger.info(f"正在删除教室管理员账号 ID: {admin_id}")
        
        # 检查管理员是否存在
        admin = ClassroomAdmin.query.get(admin_id)
        if not admin:
            logger.error(f"管理员ID {admin_id} 不存在")
            return jsonify({'error': f'管理员ID {admin_id} 不存在'}), 404
        
        # 获取相关信息用于记录
        admin_info = {
            'id': admin.id,
            'username': admin.username,
            'classroom_name': admin.classroom_name
        }
        
        # 查找关联到此教室管理员的学生
        students = User.query.filter_by(classroom_id=admin_id).all()
        student_count = len(students)
        
        # 解除学生与此教室的关联
        for student in students:
            student.classroom_id = None
            logger.info(f"解除学生 {student.name} 与教室 {admin.classroom_name} 的关联")
        
        # 删除教室管理员账号
        db.session.delete(admin)
        db.session.commit()
        
        logger.info(f"成功删除教室管理员账号: {admin_info}, 并解除了 {student_count} 名学生的关联")
        
        return jsonify({
            'success': True,
            'message': f'已删除教室管理员账号: {admin_info["classroom_name"]}',
            'admin_info': admin_info,
            'affected_students': student_count
        })
        
    except Exception as e:
        db.session.rollback()
        error_msg = str(e)
        logger.exception(f"删除教室管理员账号失败: {error_msg}")
        return jsonify({'error': f'删除失败: {error_msg}'}), 500

@app.route('/admin/sync_students_to_admins')
@admin_required
def sync_students_to_classroom_admins():
    """同步学生教室位置（规范化教室名称）"""
    try:
        # 获取所有教室管理员账号
        classroom_admins = {}
        for admin in ClassroomAdmin.query.all():
            # 用户名作为匹配键
            classroom_admins[admin.username.upper()] = admin.username
            logger.info(f"教室管理员: 用户名={admin.username}, 教室名={admin.classroom_name}")
        
        # 获取所有有教室位置的学生
        students = User.query.filter(
            User.role == 'student',
            User.classroom_location.isnot(None)
        ).all()
        
        # 关联统计
        total_students = len(students)
        updated_count = 0
        matched_count = 0
        unmatched_count = 0
        unmatched_students = []
        
        # 处理学生的教室位置
        for student in students:
            # 规范化教室位置（转为大写）
            old_location = student.classroom_location
            classroom_location = old_location.strip().upper()
            
            # 更新规范化的教室位置
            if classroom_location != old_location:
                student.classroom_location = classroom_location
                updated_count += 1
                logger.info(f"更新学生 {student.name}(ID:{student.id}) 的教室位置: {old_location} -> {classroom_location}")
            
            # 检查是否匹配教室管理员（仅用于统计）
            if classroom_location in classroom_admins:
                matched_count += 1
                logger.info(f"学生 {student.name}(ID:{student.id}) 的教室位置 {classroom_location} 匹配管理员: {classroom_admins[classroom_location]}")
            else:
                unmatched_count += 1
                unmatched_students.append({
                    "id": student.id,
                    "name": student.name,
                    "classroom_location": classroom_location
                })
                logger.warning(f"未找到匹配的教室管理员: 学生 {student.name}(ID:{student.id}), 教室位置: {classroom_location}")
        
        # 提交更改
        db.session.commit()
        
        # 返回结果
        result = {
            "total_students": total_students,
            "updated_count": updated_count,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "unmatched_students": unmatched_students,
            "available_classrooms": list(classroom_admins.keys())
        }
        
        flash(f'已规范化 {updated_count} 名学生的教室位置。{matched_count} 名学生匹配到了教室管理员，{unmatched_count} 名学生未匹配到教室管理员。', 'success' if unmatched_count == 0 else 'warning')
        return render_template('sync_result.html', result=result)
        
    except Exception as e:
        db.session.rollback()
        logger.exception(f"同步学生教室位置失败: {str(e)}")
        flash(f'同步失败: {str(e)}', 'danger')
        return redirect(url_for('student_list'))

@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(app.static_folder, filename)

if __name__ == '__main__':
    # 获取本机IP
    local_ip = get_local_ip()
    
    # 生成SSL证书
    ssl_ready = generate_ssl_cert()
    
    if ssl_ready:
        # 使用SSL证书运行Flask应用
        cert_file = 'cert.pem'
        key_file = 'key.pem'
        print("\n" + "="*50)
        print("使用HTTPS启动服务器...")
        print("请使用以下地址访问应用：")
        print(f"本机访问: https://localhost:5000")
        print(f"局域网访问 (手机等设备): https://{local_ip}:5000")
        print("首次访问时会显示安全警告，这是正常的，请点击'高级'然后'继续访问'")
        print("=" * 50 + "\n")
        app.run(debug=True, host='0.0.0.0', ssl_context=(cert_file, key_file))
    else:
        # 如果SSL证书生成失败，则以普通HTTP模式运行
        print("\n" + "="*50)
        print("SSL证书生成失败，将以HTTP模式启动服务器")
        print("注意：HTTP模式下，移动设备的摄像头功能可能无法正常工作")
        print(f"本机访问: http://localhost:5000")
        print(f"局域网访问: http://{local_ip}:5000")
        print("=" * 50 + "\n")
        app.run(debug=True, host='0.0.0.0') 